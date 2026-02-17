import openpyxl

def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value

def set_cell_data(path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(path)

import openpyxl

def get_data(url, sheet_name):
    workbook = openpyxl.load_workbook(url)
    sheet = workbook[sheet_name]

    final_list = []
    for i in range(2, sheet.max_row + 1):  # desde la fila 2
        email = sheet.cell(row=i, column=1).value   # col A
        password = sheet.cell(row=i, column=2).value  # col B

        # normaliza
        email = str(email).strip() if email is not None else None
        password = str(password).strip() if password is not None else None

        # ignora filas vacías
        if not email and not password:
            continue   # o break si quieres parar al primer vacío

        final_list.append((email, password))

    print("COUNT:", len(final_list))
    print("DATA:", final_list)
    return final_list

